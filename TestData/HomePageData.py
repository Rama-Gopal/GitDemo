import openpyxl


class HomePageData:

    test_HomePage_data = [{"firstname": "Rahul", "lastname": "shetty", "gender": "Male"}, {"firstname": "Anshika", "lastname": "shetty", "gender": "Female"}]

    @staticmethod
    def getTestData(test_case_name):
        book = openpyxl.load_workbook("//Users//peddinti.gopal//Documents//PythonDemo.xlsx")
        sheet = book.active
        Dict = {}
        for row in range(1, sheet.max_row+1):
            if sheet.cell(row=row, column=1).value == test_case_name:
                for col in range(2, sheet.max_column+1):
                    Dict[sheet.cell(row=1, column=col).value] = sheet.cell(row=row, column=col).value
                break
        return [Dict]
