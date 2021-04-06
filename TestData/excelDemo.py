import openpyxl

book = openpyxl.load_workbook("//Users//peddinti.gopal//Documents//PythonDemo.xlsx")
sheet = book.active
cell = sheet.cell(row=1, column=2)
print(cell.value)
sheet.cell(row=2, column=2).value = "Rahul"
print(sheet.cell(row=2, column=2).value)
print(sheet.max_row, sheet.max_column)
print(sheet['A5'].value)
for row in range(1, sheet.max_row+1):
    for col in range(1, sheet.max_column+1):
        print(sheet.cell(row=row, column=col).value, end=' ')
    print()

#  to get a particular row

for row in range(1, sheet.max_row+1):
    if sheet.cell(row=row, column=1).value == "Testcase2":
        for col in range(1, sheet.max_column+1):
            print(sheet.cell(row=row, column=col).value, end=" ")
        break

# building a dictionary for data driven
Dict = {}
for row in range(1, sheet.max_row+1):
    if sheet.cell(row=row, column=1).value == "Testcase2":
        for col in range(2, sheet.max_column+1):
            Dict[sheet.cell(row=1, column=col).value] = sheet.cell(row=row, column=col).value
        break
print(Dict)
